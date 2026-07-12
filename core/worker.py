import os
import csv
import re
import glob
from datetime import datetime, date
from PySide6.QtCore import QThread, Signal
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import dateutil.parser

from core.category_manager import CategoryManager, CONFIG_FILE
from core.user_keywords_manager import UserKeywordsManager


class ProcessingWorker(QThread):
    progress = Signal(int)
    status = Signal(str)
    finished = Signal(dict)
    error = Signal(str)

    def __init__(self, file_paths):
        super().__init__()
        self.file_paths = file_paths
        self.manager = CategoryManager()
        self.keywords_manager = UserKeywordsManager()

    def run(self):
        try:
            self.manager.load()
            self.keywords_manager.load()
            total_files = len(self.file_paths)
            all_data = []
            unclassified_list = []

            for idx, file_path in enumerate(self.file_paths):
                self.status.emit(f"处理: {os.path.basename(file_path)}")
                if os.path.isdir(file_path):
                    self._process_folder(file_path, all_data, unclassified_list)
                else:
                    self._process_file(file_path, all_data, unclassified_list)
                self.progress.emit(int((idx + 1) / total_files * 100))

            self.status.emit("生成输出文件...")
            output_file, unclassified_file = self._generate_output(all_data, unclassified_list)

            self.finished.emit({
                'total': len(all_data),
                'unclassified': len(unclassified_list),
                'output_file': output_file,
                'unclassified_file': unclassified_file
            })

        except Exception as e:
            self.error.emit(str(e))

    def _process_folder(self, folder_path, all_data, unclassified_list):
        for ext in ['*.csv', '*.xlsx', '*.xls']:
            for file_path in glob.glob(os.path.join(folder_path, ext)):
                self._process_file(file_path, all_data, unclassified_list)

    def _read_file(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            for encoding in ['utf-8', 'ansi', 'gbk']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        return list(csv.reader(f))
                except:
                    continue
            return None
        elif ext in ['.xlsx', '.xls']:
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                rows = []
                for row in ws.iter_rows(values_only=True):
                    processed = []
                    for cell in row:
                        if cell is None:
                            processed.append('')
                        elif isinstance(cell, (datetime, date)):
                            processed.append(cell.strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            processed.append(str(cell))
                    rows.append(processed)
                wb.close()
                return rows
            except Exception as e:
                print(f"读取Excel失败: {e}")
                return None
        return None

    def _detect_file_type(self, header_row):
        wechat_headers = ['交易时间', '交易类型', '交易对方', '商品', '收/支', '金额(元)']
        alipay_headers = ['交易时间', '交易分类', '交易对方', '对方账号', '商品说明', '收/支']
        
        # 检查微信
        wechat_match = sum(1 for h in wechat_headers if h in header_row)
        if wechat_match >= 3:
            return 'wechat'
        
        # 检查支付宝
        alipay_match = sum(1 for h in alipay_headers if h in header_row)
        if alipay_match >= 3:
            return 'alipay'
        
        return None

    def _find_header_row(self, data_rows):
        """查找表头行"""
        wechat_keywords = ['交易时间', '交易类型', '交易对方', '商品']
        alipay_keywords = ['交易时间', '交易分类', '交易对方', '商品说明']
        
        for idx, row in enumerate(data_rows):
            row_str = [str(cell).strip() for cell in row if cell]
            # 检查微信
            if all(k in row_str for k in wechat_keywords[:3]):
                return idx, 'wechat'
            # 检查支付宝
            if all(k in row_str for k in alipay_keywords[:3]):
                return idx, 'alipay'
        
        return -1, None

    def _process_file(self, file_path, all_data, unclassified_list):
        data_rows = self._read_file(file_path)
        if not data_rows:
            print(f"  跳过: 无法读取文件 {file_path}")
            return

        print(f"  共 {len(data_rows)} 行数据")

        # 查找表头行
        file_type = None
        header_row_index = 0

        for idx, row in enumerate(data_rows):
            row_str = [str(cell).strip() for cell in row if cell]
            # 微信表头
            if '交易时间' in row_str and '交易类型' in row_str and '商品' in row_str:
                file_type = 'wechat'
                header_row_index = idx
                break
            # 支付宝表头
            if '交易时间' in row_str and '交易分类' in row_str and '商品说明' in row_str:
                file_type = 'alipay'
                header_row_index = idx
                break

        if not file_type:
            print(f"  跳过: 无法识别文件格式")
            return

        print(f"  识别为: {file_type}")
        print(f"  找到表头在第 {header_row_index + 1} 行")

        # 从表头下一行开始
        data_rows = data_rows[header_row_index + 1:]

        print(f"  开始处理 {len(data_rows)} 行数据...")

        processed_count = 0
        unclassified_count = 0

        for row_idx, row in enumerate(data_rows, 1):
            if len(row) < 6:
                continue

            try:
                # 解析日期
                date_str = str(row[0]).strip() if row[0] else ''
                if not date_str:
                    continue
                date_obj = dateutil.parser.parse(date_str)

                # ========== 微信 ==========
                if file_type == 'wechat':
                    in_out = str(row[4]).strip() if len(row) > 4 else ''
                    store = str(row[2]).strip() if len(row) > 2 else ''
                    commodity = str(row[3]).strip() if len(row) > 3 else ''
                    money_str = str(row[5]).strip() if len(row) > 5 else ''
                    payment_raw = str(row[6]).strip() if len(row) > 6 else ''
                    trans_type = str(row[1]).strip() if len(row) > 1 else ''
                    search_text = f"{store} {commodity}"
                    if payment_raw in ["零钱", "/", "", "None"]:
                        payment = "微信钱包"
                    else:
                        payment = payment_raw

                # ========== 支付宝 ==========
                else:
                    in_out = str(row[5]).strip() if len(row) > 5 else ''
                    
                    # 先判断不计收支，直接跳过
                    if in_out == "不计收支" or not in_out:
                        continue
                    
                    store = str(row[2]).strip() if len(row) > 2 else ''
                    commodity = str(row[4]).strip() if len(row) > 4 else ''
                    money_str = str(row[6]).strip() if len(row) > 6 else ''
                    payment_raw = str(row[7]).strip() if len(row) > 7 else ''
                    trans_type = str(row[1]).strip() if len(row) > 1 else ''
                    search_text = f"{commodity} {store}"

                    # 统一支付宝支付方式
                    if "账户余额" in payment_raw:
                        payment = "支付宝"
                    elif "花呗" in payment_raw:
                        payment = "花呗"
                    else:
                        import re
                        match = re.search(r'([\u4e00-\u9fa5]+(?:储蓄卡|信用卡))\((\d+)\)', payment_raw)
                        if match:
                            payment = f"{match.group(1)}({match.group(2)})"
                        else:
                            payment = payment_raw

                # 解析金额
                money_match = re.search(r"-?\d+\.?\d*", money_str)
                if not money_match:
                    continue
                money = float(money_match.group(0))

                cat1, cat2, cat3, cat4 = self._classify(in_out, search_text)

                data = {
                    'date': date_obj,
                    'in_out': in_out,
                    'money': money,
                    'store': store,
                    'commodity': commodity,
                    'payment': payment,
                    'trans_type': trans_type,
                    'cat1': cat1,
                    'cat2': cat2,
                    'cat3': cat3,
                    'cat4': cat4
                }

                if cat1 is None or cat2 is None or cat3 is None:
                    unclassified_list.append(data)
                    unclassified_count += 1
                else:
                    all_data.append(data)
                    processed_count += 1

            except Exception as e:
                print(f"  处理第 {row_idx} 行时出错: {e}")
                continue

        print(f"  成功处理 {processed_count} 笔交易，{unclassified_count} 笔未分类")
        
    def _match_rules(self, text, rules_str):
        """匹配规则"""
        if not rules_str:
            return False
        if not isinstance(rules_str, str):
            return False
        patterns = rules_str.split('|')
        for pattern in patterns:
            pattern = pattern.strip()
            if not pattern:
                continue
            try:
                if re.search(pattern, text, re.IGNORECASE):
                    return True
            except:
                if pattern.lower() in text.lower():
                    return True
        return False


    def _classify(self, in_out, search_text):
        # 先匹配用户关键词
        for path, keywords in self.keywords_manager.get_all_keywords().items():
            if not keywords:
                continue
            parts = path.split('/')
            if len(parts) >= 3:
                main_cat = parts[0]
                sub_cat = parts[1]
                third_cat = parts[2]
                fourth_cat = parts[3] if len(parts) > 3 else None
                for keyword in keywords:
                    try:
                        if re.search(keyword, search_text, re.IGNORECASE):
                            print(f"  ✅ 用户关键词匹配: '{keyword}' → {path}")
                            return main_cat, sub_cat, third_cat, fourth_cat
                    except:
                        if keyword.lower() in search_text.lower():
                            print(f"  ✅ 用户关键词匹配: '{keyword}' → {path}")
                            return main_cat, sub_cat, third_cat, fourth_cat

        # 再匹配系统规则
        for main_cat, sub_cats in self.manager.categories.items():
            if not isinstance(sub_cats, dict):
                continue
            for sub_cat, third_cats in sub_cats.items():
                if not isinstance(third_cats, dict):
                    continue
                for third_cat, rules in third_cats.items():
                    if rules and isinstance(rules, str) and self._match_rules(search_text, rules):
                        return main_cat, sub_cat, third_cat, None

        return None, None, None, None
        
        def _match_rules(self, text, rules_str):
            if not rules_str:
                return False
            if not isinstance(rules_str, str):
                return False
            patterns = rules_str.split('|')
            for pattern in patterns:
                pattern = pattern.strip()
                if not pattern:
                    continue
                try:
                    if re.search(pattern, text, re.IGNORECASE):
                        return True
                except:
                    if pattern.lower() in text.lower():
                        return True
            return False

    def _generate_output(self, all_data, unclassified_list):
        if not self.file_paths:
            output_dir = "完成"
        else:
            output_dir = os.path.join(os.path.dirname(self.file_paths[0]), '完成')
        os.makedirs(output_dir, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "所有交易"

        headers = ['日期', '收支类型', '金额', '类别', '子类', '所属账本', '收支账户', '备注']
        for col, header in enumerate(headers, 1):
            ws.cell(1, col).value = header
            ws.cell(1, col).font = Font(bold=True)

        for idx, data in enumerate(all_data, 2):
            ws.cell(idx, 1).value = data['date']
            ws.cell(idx, 2).value = data['in_out']
            ws.cell(idx, 3).value = data['money']
            ws.cell(idx, 4).value = data['cat2'] if data['cat2'] else ''
            ws.cell(idx, 5).value = data['cat3'] if data['cat3'] else ''
            ws.cell(idx, 6).value = "日常账本"
            ws.cell(idx, 7).value = data['payment']
            ws.cell(idx, 8).value = f"{data['commodity']}-{data['store']}"

        output_file = os.path.join(output_dir, "sc.xlsx")
        counter = 1
        while os.path.exists(output_file):
            try:
                with open(output_file, 'a') as f:
                    pass
                break
            except:
                output_file = os.path.join(output_dir, f"sc_{counter}.xlsx")
                counter += 1
        wb.save(output_file)

        unclassified_file = None
        if unclassified_list:
            unclassified_file = os.path.join(output_dir, "未分类交易.xlsx")
            counter = 1
            while os.path.exists(unclassified_file):
                try:
                    with open(unclassified_file, 'a') as f:
                        pass
                    break
                except:
                    unclassified_file = os.path.join(output_dir, f"未分类交易_{counter}.xlsx")
                    counter += 1

            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = "未分类交易"
            headers2 = ['日期', '收支类型', '金额', '类别', '子类', '所属账本', '收支账户', '备注']
            for col, header in enumerate(headers2, 1):
                ws2.cell(1, col).value = header
                ws2.cell(1, col).font = Font(bold=True)
            for idx, data in enumerate(unclassified_list, 2):
                ws2.cell(idx, 1).value = data['date']
                ws2.cell(idx, 2).value = data['in_out']
                ws2.cell(idx, 3).value = data['money']
                ws2.cell(idx, 4).value = ''
                ws2.cell(idx, 5).value = ''
                ws2.cell(idx, 6).value = "日常账本"
                ws2.cell(idx, 7).value = data['payment']
                ws2.cell(idx, 8).value = f"{data['commodity']}-{data['store']}"
            wb2.save(unclassified_file)

        return output_file, unclassified_file
