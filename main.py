import csv
import glob
import tkinter as tk
import dateutil.parser
import re
import os
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime, date
from classification import classify_transaction  # 改名

# 常量定义
WECHAT_HEADER = ['交易时间', '交易类型', '交易对方', '商品', '收/支', '金额(元)', 
                 '支付方式', '当前状态', '交易单号', '商户单号', '备注']
ALIPAY_HEADER = ['交易时间', '交易分类', '交易对方', '对方账号', '商品说明', 
                 '收/支', '金额', '收/付款方式', '交易状态', '交易订单号', 
                 '商家订单号', '备注', '']

def read_file_content(file_path):
    """智能读取文件，支持 csv 和 xlsx 格式"""
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.csv':
        return read_csv_file(file_path)
    elif file_ext in ['.xlsx', '.xls']:
        return read_excel_file(file_path)
    else:
        print(f"不支持的文件格式: {file_ext}")
        return None

def read_csv_file(file_path):
    """读取CSV文件"""
    data_rows = []
    for encoding in ['utf-8', 'ansi', 'gbk']:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                data_rows = list(reader)
                break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"读取CSV失败 {file_path}: {e}")
            return None
    
    return data_rows if data_rows else None

def read_excel_file(file_path):
    """读取Excel文件（xlsx/xls格式）"""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        data_rows = []
        for row in ws.iter_rows(values_only=True):
            processed_row = []
            for cell in row:
                if cell is None:
                    processed_row.append('')
                elif isinstance(cell, (datetime, date)):
                    processed_row.append(cell.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    processed_row.append(str(cell))
            data_rows.append(processed_row)
        
        wb.close()
        return data_rows
    except Exception as e:
        print(f"读取Excel失败 {file_path}: {e}")
        return None

def detect_file_type(header_row):
    """检测文件类型（微信/支付宝）"""
    if header_row == WECHAT_HEADER:
        return 'wechat'
    elif header_row == ALIPAY_HEADER:
        return 'alipay'
    
    # 尝试部分匹配
    wechat_match = sum(1 for h in WECHAT_HEADER[:5] if h in header_row)
    alipay_match = sum(1 for h in ALIPAY_HEADER[:5] if h in header_row)
    
    if wechat_match > alipay_match and wechat_match >= 3:
        return 'wechat'
    elif alipay_match > wechat_match and alipay_match >= 3:
        return 'alipay'
    
    return None

def export_unclassified_to_excel(unclassified_list, output_path):
    """导出未分类的交易到单独的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "未分类交易"
    
    # 设置表头
    headers = ['序号', '日期', '收支类型', '金额', '商家名称', '商品名称', '交易类型', '支付方式', '备注']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True)
    
    # 写入数据
    for idx, item in enumerate(unclassified_list, 1):
        row_num = idx + 1
        ws.cell(row_num, 1).value = idx
        ws.cell(row_num, 2).value = item.get('date', '')
        ws.cell(row_num, 3).value = item.get('in_out', '')
        ws.cell(row_num, 4).value = item.get('money', 0)
        ws.cell(row_num, 5).value = item.get('store', '')
        ws.cell(row_num, 6).value = item.get('commodity', '')
        ws.cell(row_num, 7).value = item.get('transaction_type', '')
        ws.cell(row_num, 8).value = item.get('payment', '')
        ws.cell(row_num, 9).value = item.get('note', '')
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"未分类交易已保存到: {output_path}")

def show_unclassified_warning(unclassified_list, output_path):
    """显示未分类交易的警告框"""
    count = len(unclassified_list)
    
    # 构建警告消息
    message = f"发现 {count} 笔未分类的交易！\n\n"
    message += f"未分类交易列表已保存到:\n{output_path}\n\n"
    message += "前10笔未分类交易：\n"
    message += "-" * 50 + "\n"
    
    # 显示前10笔
    for i, item in enumerate(unclassified_list[:10], 1):
        date_str = item.get('date', '')
        if isinstance(date_str, datetime):
            date_str = date_str.strftime('%Y-%m-%d')
        store = item.get('store', '')
        money = item.get('money', 0)
        message += f"{i}. {date_str} | {store} | {money}元\n"
    
    if count > 10:
        message += f"\n... 还有 {count - 10} 笔未显示"
    
    message += "\n\n请检查并更新分类规则！"
    
    # 弹出警告框
    root = tk.Tk()
    root.withdraw()
    messagebox.showwarning("未分类交易警告", message)
    root.destroy()

# 主程序
root = tk.Tk()
root.withdraw()
folder_path = filedialog.askdirectory()

if not folder_path:
    print("未选择文件夹")
    exit()

# 删除旧的临时文件
path1 = os.path.join(folder_path, "DelLoad.csv")
try:
    os.remove(path1)
except FileNotFoundError:
    pass
except Exception as e:
    print(f"删除临时文件失败: {e}")

# 获取所有支持的文件
supported_files = []
supported_files.extend(glob.glob(os.path.join(folder_path, "*.csv")))
supported_files.extend(glob.glob(os.path.join(folder_path, "*.xlsx")))
supported_files.extend(glob.glob(os.path.join(folder_path, "*.xls")))

if not supported_files:
    print("未找到支持的文件（CSV/XLSX/XLS）")
    exit()

print(f"找到 {len(supported_files)} 个文件待处理")

# 初始化数据存储
all_dates = []
all_in_out = []
all_stores = []
all_category1 = []
all_category2 = []
all_money = []
all_payment = []
all_commodity = []
all_transaction_type = []

# 用于存储未分类的交易
unclassified_transactions = []

for file_path in supported_files:
    print(f"\n处理文件: {os.path.basename(file_path)}")
    
    # 读取文件内容
    data_rows = read_file_content(file_path)
    if not data_rows:
        print(f"  跳过: 无法读取文件")
        continue
    
    print(f"  共 {len(data_rows)} 行数据")
    
    # 检测文件类型
    file_type = None
    header_row_index = 0
    
    # 先检查第一行
    file_type = detect_file_type(data_rows[0])
    if not file_type:
        # 如果第一行不是表头，尝试查找
        for idx, row in enumerate(data_rows):
            if '交易时间' in row and '交易对方' in row:
                file_type = 'wechat' if '商品' in row else 'alipay'
                header_row_index = idx
                break
    
    if not file_type:
        print(f"  跳过: 无法识别文件格式")
        continue
    
    print(f"  识别为: {file_type}")
    
    # 删除表头之前的行
    if header_row_index > 0:
        data_rows = data_rows[header_row_index:]
        print(f"  找到表头在第 {header_row_index + 1} 行")
    
    # 处理数据
    for row_idx, row in enumerate(data_rows[1:], 1):  # 跳过表头
        if len(row) < 5:
            continue
        
        try:
            # 解析日期
            date_str = str(row[0]).strip() if row[0] else ''
            if not date_str:
                continue
            
            # 如果已经是datetime对象，直接使用
            if isinstance(row[0], (datetime, date)):
                date_obj = row[0]
            else:
                date_obj = dateutil.parser.parse(date_str)
            
            # 提取基本信息
            in_out = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            if in_out == "不计收支":
                continue
            
            store = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            commodity = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            money_str = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            payment = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            trans_type = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            
            # 解析金额
            money_match = re.search(r"-?\d+\.?\d*", money_str)
            if not money_match:
                continue
            money = float(money_match.group(0))
            
            # 标准化支付方式
            if file_type == 'wechat':
                if payment in ["零钱", "/", "", "None"]:
                    payment = "微信钱包"
            else:  # alipay
                if payment in ["/", "", "None"]:
                    payment = "支付宝"
            
            # ========== 关键修改：所有判断都在 main.py 中 ==========
            # 根据文件类型组合搜索文本
            is_unclassified = False
            if file_type == 'wechat':
                # 微信：store（商家名称）+ commodity（商品名称）
                search_text = f"{store} {commodity}"
                result = classify_transaction(in_out, search_text)
            else:  # alipay
                # 支付宝：commodity（商品说明）+ store（交易对方）
                search_text = f"{commodity} {store}"
                result = classify_transaction(in_out, search_text)

            if result:
                cat1, cat2 = result
                if cat1 == "未分类" or cat1 is None:
                    is_unclassified = True
            else:
                cat1, cat2 = None, None
                is_unclassified = True

            # ========== 关键修改：只有分类成功的才加入主列表 ==========
            if is_unclassified:
                # 未分类：只记录到未分类列表，不加入主数据
                unclassified_transactions.append({
                    'date': date_obj,
                    'in_out': in_out,
                    'money': money,
                    'store': store,
                    'commodity': commodity,
                    'transaction_type': trans_type,
                    'payment': payment,
                    'note': f"{commodity}-{store}" if commodity else store
                })
                # 注意：这里不添加到 all_* 列表
            else:
                # 分类成功：加入主数据
                all_dates.append(date_obj)
                all_in_out.append(in_out)
                all_stores.append(store)
                all_commodity.append(commodity)
                all_money.append(money)
                all_payment.append(payment)
                all_transaction_type.append(trans_type)
                all_category1.append(cat1 if cat1 else "未分类")
                all_category2.append(cat2 if cat2 else "未分类")
            
        except Exception as e:
            print(f"  处理第 {row_idx} 行时出错: {e}")
            continue

print(f"\n总计处理 {len(all_dates)} 笔交易")
print(f"其中未分类: {len(unclassified_transactions)} 笔")

if len(all_dates) == 0:
    print("没有成功处理任何交易，请检查文件格式")
    exit()

# 创建主Excel输出
wb = Workbook()
ws = wb.active
ws.title = "所有交易"

# 设置表头
headers = ['日期', '收支类型', '金额', '类别', '子类', '所属账本', '收支账户', '备注']
for col, header in enumerate(headers, 1):
    ws.cell(1, col).value = header

# 写入数据
for idx in range(len(all_dates)):
    row_num = idx + 2
    
    ws.cell(row_num, 1).value = all_dates[idx]
    ws.cell(row_num, 2).value = all_in_out[idx]
    ws.cell(row_num, 3).value = all_money[idx]
    
    # 处理大餐分类
    cat2 = all_category2[idx]
    if cat2 == "三餐" and all_money[idx] > 20:
        cat2 = "大餐"
    
    ws.cell(row_num, 4).value = all_category1[idx]
    ws.cell(row_num, 5).value = cat2
    ws.cell(row_num, 6).value = "日常账本"
    ws.cell(row_num, 7).value = all_payment[idx]
    ws.cell(row_num, 8).value = f"{all_commodity[idx]}-{all_stores[idx]}"

# 保存主文件
output_dir = os.path.join(folder_path, "完成")
os.makedirs(output_dir, exist_ok=True)

output_file = os.path.join(output_dir, "sc.xlsx")
wb.save(output_file)
print(f"\n主文件已保存到: {output_file}")

# 如果有未分类的交易，导出额外的Excel并弹出警告
if unclassified_transactions:
    # 导出未分类交易
    unclassified_file = os.path.join(output_dir, "未分类交易.xlsx")
    export_unclassified_to_excel(unclassified_transactions, unclassified_file)
    
    # 显示警告框
    show_unclassified_warning(unclassified_transactions, unclassified_file)
    
    # 在控制台也显示摘要
    print(f"\n⚠️ 发现 {len(unclassified_transactions)} 笔未分类的交易")
    print(f"详细列表已保存到: {unclassified_file}")
    print("\n前5笔未分类交易:")
    for i, item in enumerate(unclassified_transactions[:5], 1):
        date_str = item['date']
        if isinstance(date_str, datetime):
            date_str = date_str.strftime('%Y-%m-%d')
        print(f"  {i}. {date_str} | {item['store']} | {item['money']}元")
    if len(unclassified_transactions) > 5:
        print(f"  ... 还有 {len(unclassified_transactions) - 5} 笔")
else:
    print("\n✅ 所有交易已成功分类！")

# 打开输出文件夹
os.startfile(output_dir)